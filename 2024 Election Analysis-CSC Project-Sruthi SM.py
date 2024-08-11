import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
#INDIA-Alliance
#visual formatting
font = {'size':9,
        'family': 'Ink Free',
        'weight': 'bold'}
plt.rc('font', **font)
Explode=[0.005,0.1,0.07,0.05,0.02,0.01,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0, 0]
Colours1= [
    "#556B2F", "#808000", "#6B8E23", "#9ACD32", "#8FBC8F",
    "#2E8B57", "#3CB371", "#66CDAA", "#5F9EA0", "#20B2AA",
    "#BDB76B", "#F0E68C", "#EEE8AA", "#DAA520", "#FFD700",
    "#FF8C00", "#D2B48C", "#F4A460", "#CD853F", "#D2691E",
    "#8B4513", "#A0522D", "#E9967A", "#FF7F50", "#F08080"
]
header_fill = PatternFill(start_color="6B8E23", end_color="6B8E23", fill_type="solid")  # Olive Drab
row_fill1 = PatternFill(start_color="8A9A5B", end_color="8A9A5B", fill_type="solid")  # Moss Green
row_fill2 = PatternFill(start_color="A3A57A", end_color="A3A57A", fill_type="solid")  # Pale Olive Green

#data
Parties_INDIA = [
    "Indian National Congress", "Samajwadi Party", "All India Trinamool Congress",
    "Dravida Munnetra Kazhagam", "Communist Party of India (Marxist)", "Rashtriya Janata Dal",
    "Shiv Sena (Uddhav Balasaheb Thackeray)", "Aam Aadmi Party", "Nationalist Congress Party (Sharadchandra Pawar)",
    "Communist Party of India", "Jharkhand Mukti Morcha", "Communist Party of India (Marxist-Leninist) Liberation",
    "Indian Union Muslim League", "Jammu and Kashmir National Conference", "Viduthalai Chiruthaigal Katchi",
    "Bharat Adivasi Party", "Kerala Congress", "Marumalarchi Dravida Munnetra Kazhagam",
    "Rashtriya Loktantrik Party", "Revolutionary Socialist Party", "All India Forward Bloc",
    "Jammu and Kashmir Peoples Democratic Party", "Vikassheel Insaan Party", "Assam Jatiya Parishad",
    "Kerala Congress (Mani)"
]
Votes_INDIA = [
    136759064, 29549381, 28213393, 11754710, 11342553,
    10107402, 9567779, 7147800, 5921162, 3132683,
    2627488, 1726309, 1199839, 1139084, 990237,
    1257056, 364631, 542213, 596955, 587303,
    289941, 435980, 1187455, 414441, 277365
]
#main code
#finding sum
sum=0
count=0
for i in Votes_INDIA:
    sum+=Votes_INDIA[count]
    count+=1
count=0
Percentage_INDIA=[]
for j in Votes_INDIA:
    Per=str(round((Votes_INDIA[count]/sum)*100,2))
    count+=1
    Percentage_INDIA.append(Per+"%")
#new list for legend
count=0
newlist=[]
for k in Parties_INDIA:
    l=k+' : '+Percentage_INDIA[count]+"%"
    count+=1
    newlist.append(l)

INDIA_dict={"Party":Parties_INDIA,"Votes":Votes_INDIA,"Percentage in Alliance":Percentage_INDIA}
INDIA_df=pd.DataFrame(INDIA_dict)
INDIA_df.to_excel("INDIA.xlsx", index=False)
workbook = load_workbook("INDIA.xlsx")
sheet = workbook.active
sheet.column_dimensions['A'].width = 45
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 22
for cell in sheet[1]:  # Assuming headers are in the first row
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")  # White font for contrast
for row_index, row in enumerate(sheet.iter_rows(min_row=2)):  # Start from the second row
    for cell in row:
        fill_color = row_fill1 if row_index % 2 == 0 else row_fill2
        cell.fill = fill_color
workbook.save("INDIA.xlsx")

fig, ax = plt.subplots()
ax.set_title("Vote Percentage of Parties in 'INDIA' with respect to Total Votes of 'INDIA'",font="Impact",fontsize=18,loc='right',pad=20,x=0.42,y=0.97,color="white")

fig.set_figheight(8)
fig.set_figwidth(16)
fig.patch.set_facecolor('#7D9A5C') 
ax.set_facecolor('#7D9A5C')

ax.pie(Votes_INDIA, startangle=60, explode=Explode, radius=1.5, colors=Colours1)
ax.legend(newlist, loc='upper right', fontsize=9.7, bbox_to_anchor=(0.6, 1), ncol=2) 
plt.savefig("Vote Percentage of Parties in 'INDIA' with respect to Total Votes of 'INDIA'.png")
plt.show()
#NDA-ALLIANCE
#visual formatting
font = {'size':9,
        'family': 'Ink Free',
        'weight': 'bold'}
plt.rc('font', **font)
Explode=[0.005,0.15,0.1,0.07,0.05,0.01,0.005,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
Colours2 = [
    "#003366", "#003399", "#004080", "#00509e", "#0066cc",
    "#0073e6", "#3399ff", "#66b3ff", "#99c2ff", "#b79fcd",
    "#c1a3e0", "#d3a6c2", "#d3a1e5", "#d39bd4", "#b993d3",
    "#9b5de5", "#8a2b9e", "#7a5b9e", "#6a5acd", "#9c6c9d",
    "#8a2d77", "#7b4d8c", "#6d28d9", "#5a2d81", "#4b0082",
    "#6a0dad", "#8a2d9d", "#9b2d77"
]
header_fill = PatternFill(start_color="1A1A68", end_color="1A1A68", fill_type="solid")  # Navy blue
row_fill1 = PatternFill(start_color="3B3C8A", end_color="3B3C8A", fill_type="solid")  # Deep blue-purple
row_fill2 = PatternFill(start_color="5F5FBB", end_color="5F5FBB", fill_type="solid")  # Slightly lighter blue-purple
#data
Parties_NDA = [
    "Bharatiya Janata Party", "Telugu Desam Party", "Janata Dal (United)",
    "Shiv Sena", "Lok Janshakti Party (Ram Vilas)", "Janata Dal (Secular)",
    "Jana Sena Party", "Rashtriya Lok Dal", "All Jharkhand Students Union",
    "Nationalist Congress Party", "United People's Party Liberal", "Sikkim Krantikari Morcha",
    "Apna Dal (Soneylal)", "Asom Gana Parishad", "Hindustani Awam Morcha",
    "Pattali Makkal Katchi", "Bharath Dharma Jana Sena", "Tamil Maanila Congress",
    "Amma Makkal Munnettra Kazhagam", "National People's Party", "Naga People's Front",
    "Nationalist Democratic Progressive Party", "Rashtriya Lok Morcha", "Rashtriya Samaj Paksha",
    "Suheldev Bharatiya Samaj Party", "Independent politician"
]
Votes_NDA = [
    235973935, 12775270, 8039663, 7401447, 2810250,
    2173701, 1454158, 893460, 458677, 2059179,
    488995, 164396, 808245, 1298707, 494960,
    1879689, 505753, 410401, 393415, 417930,
    299536, 350967, 253876, 521246, 340188,
    342882
]
#main Code
#finding sum
count=0
sum=0
for i in Votes_NDA:
    sum+=Votes_NDA[count]
    count+=1
count=0
Percentage_NDA=[]
for j in Votes_NDA:
    Per=str(round((Votes_NDA[count]/sum)*100,2))
    count+=1
    Percentage_NDA.append(Per+"%")
pd.Series(Percentage_NDA)
#new list for legend
count=0
newlist=[]
for k in Parties_NDA:
    l=k+' : '+Percentage_NDA[count]
    count+=1
    newlist.append(l)

NDA_dict={"Party":Parties_NDA,"Votes":Votes_NDA,"Percentage in Alliance":Percentage_NDA}
NDA_df=pd.DataFrame(NDA_dict)
NDA_df.to_excel("NDA.xlsx", index=False)
workbook = load_workbook("NDA.xlsx")
sheet = workbook.active
sheet.column_dimensions['A'].width = 45
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 22
for cell in sheet[1]:  # Assuming headers are in the first row
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")  # White font for contrast
for row_index, row in enumerate(sheet.iter_rows(min_row=2)):  # Start from the second row
    for cell in row:
        fill_color = row_fill1 if row_index % 2 == 0 else row_fill2
        cell.fill = fill_color
workbook.save("NDA.xlsx")

fig, ax = plt.subplots()
ax.set_title("Vote Percentage of Parties in 'NDA' with respect to Total Votes of 'NDA'",font="Impact",fontsize=18,loc='right',pad=20,x=0.487,y=0.97,color="white")
fig.set_figheight(8)
fig.set_figwidth(16)

fig.patch.set_facecolor('#6E6E6E')
ax.set_facecolor('#6E6E6E')

ax.pie(Votes_NDA, startangle=10, explode=Explode, radius=1.5, colors=Colours2)
ax.legend(newlist, loc='upper right', fontsize=10.5, bbox_to_anchor=(0.5, 1), ncol=2) 

plt.savefig("Vote Percentage of Parties in 'NDA' with respect to Total Votes of 'NDA'.png")
plt.show()

#visual formatting
font = {'size' : 7,
        'family': 'Ink Free',
        'weight': 'bold'}
plt.rc('font', **font)
Colors3 = [
    "#5D432C", "#624932", "#674E39", "#6C5440", "#715947",
    "#765E4E", "#7B6455", "#806A5C", "#857063", "#8A766A",
    "#8F7C71", "#948278", "#9A8D81", "#9FB391", "#93A098",
    "#87AC9F", "#7BB7A7", "#6FC2AE", "#63CDC5", "#57D7CC",
    "#4BE1D3", "#3FEBDC", "#33F5E4", "#27FFEB", "#1FFFF1"
]
Dotsize=[]
Labels=[283010926,267132224,13316039,1814318,1400215,208552,691820,571078,13153818,9413379,8952587,3657237,625954,521749,140264,226975,113827,1128616,64578,77171,777570,44563,7726712,17792407,6372220]
for i in Labels:
    i//=100000
    Dotsize.append(i)
header_fill = PatternFill(start_color="4E3A2F", end_color="4E3A2F", fill_type="solid")  # Deep earthy brown
row_fill1 = PatternFill(start_color="8C6A53", end_color="8C6A53", fill_type="solid")  # Strong beige
row_fill2 = PatternFill(start_color="9F8B6B", end_color="9F8B6B", fill_type="solid")  # Rich greenish beige

#data
Alliances=["National Democratic Alliance","Indian National Developmental\nInclusive Alliance","YSR Congress Party","Shiromani Akali Dal","All India Majlis-e-Ittehadul\nMuslimeen","Zoram People's Movement","Azad Samaj Party","Voice of the People Party","Bahujan Samaj Party","Biju Janata Dal","All India Anna Dravida\nMunnetra Kazhagam","Bharat Rashtra Samithi","All India United Democratic Front","Shiromani Akali Dal (Amritsar)","Mizo National Front","Indian National Lok Dal","Jannayak Janta Party","Desiya Murpokku Dravida\nKazhagam","Revolutionary Goans Party","Sikkim Democratic Front","Bodoland People's Front","United Democratic Party","Other Parties","Independents","NOTA"]
Votes_Alliances=[283010926,267132224,13316039,1814318,1400215,208552,691820,571078,13153818,9413379,8952587,3657237,625954,521749,140264,226975,113827,1128616,64578,77171,777570,44563,7726712,17792407,6372220]

#main code
#finding sum
sum=0
count=0
for i in Votes_Alliances:
    sum+=Votes_Alliances[count]
    count+=1
count=0
Percentage_Alliances=[]
for j in Votes_Alliances:
    Per=str(round((Votes_Alliances[count]/sum)*100,2))
    count+=1
    Percentage_Alliances.append(Per+"%")

Alliances_dict={"Alliance":Alliances,"Votes":Votes_Alliances,"Percentage in Total votes":Percentage_Alliances}
Alliances_df=pd.DataFrame(Alliances_dict)
Alliances_df.to_excel("Alliance.xlsx", index=False)
workbook = load_workbook("Alliance.xlsx")
sheet = workbook.active
sheet.column_dimensions['A'].width = 45
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 26
for cell in sheet[1]:  # Assuming headers are in the first row
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")  # White font for contrast
for row_index, row in enumerate(sheet.iter_rows(min_row=2)):  # Start from the second row
    for cell in row:
        fill_color = row_fill1 if row_index % 2 == 0 else row_fill2
        cell.fill = fill_color
workbook.save("Alliance.xlsx")

fig, ax = plt.subplots()
ax.set_title("Vote Percentage of Alliances with respect to Total Votes",font="Impact",fontsize=15,loc='left',color="white")
ax.set_xlabel("Votes secured in Hundred Millions",color="white",fontsize=12)
ax.set_ylabel("Alliances",color="white",fontsize=12)
fig.set_figheight(8)
fig.set_figwidth(16)

fig.patch.set_facecolor('#261A12') 
ax.set_facecolor('#261A12')
ax.grid(color='#6C5440')
ax.tick_params(axis='x', colors='white')
ax.tick_params(axis='y', colors='white') 
ax.spines['top'].set_color('white')
ax.spines['right'].set_color('white')
ax.spines['bottom'].set_color('white')
ax.spines['left'].set_color('white')

scatter = plt.scatter(Votes_Alliances,Alliances,s=Dotsize,color=Colors3)
for i in range(len(Votes_Alliances)):
    plt.annotate(Labels[i],(Votes_Alliances[i],Alliances[i]),textcoords="offset points",xytext=(25, -2),ha='center',fontsize=6,color='white')

plt.savefig("Vote Percentage of Alliances with respect to Total Votes.png")
plt.show()

#visual formatting
font = {'size':10,
        'family': 'Ink Free',
        'weight': 'bold'}
plt.rc('font', **font)
Labels=['Seat Share Percentage between Major Parties in 2019','Seat Share Percentage between Major Parties in 2024']
header_fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")  # Teal
row_fill1 = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
row_fill2 = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # Dark Gray

#data
Major_Parties=["Bharatiya Janata Party","Indian National Congress","Samajwadi Party","All India Trinamool Congress","Dravida Munnetra Kazhagam","Telugu Desam Party","Janata Dal","Shiv Sena(UBT)","Nationalist Congress Party(SP)","Shiv Sena","Others"]
MP=["Bharatiya\nJanata\nParty","Indian\nNational\nCongress","Samajwadi\nParty","All India\nTrinamool\nCongress","Dravida\nMunnetra\nKazhagam","Telugu\nDesam\nParty","Janata\nDal","Shiv Sena\n(UBT)","Nationalist\nCongress\nParty(SP)","Shiv\nSena","Others"]
Seat_Percentage_2019 = [55.8, 9.6, 0.9, 4.1, 4.2, 0.6, 2.9, 3.3, 9.9, 3.3, 9.4]
Seat_Percentage_2024 = [44.20, 18.24, 6.82, 5.35, 4.06, 2.21, 2.95, 1.66, 1.48, 1.29, 11.74]
S2019=[]
S2024=[]
#main code
Difference=[]
for i in range(len(MP)):
    Diff=str(round(Seat_Percentage_2024[i]-Seat_Percentage_2019[i],2))
    Difference.append(Diff+"%")
for i in range(len(Seat_Percentage_2019)):
    Seat_Percent=str(Seat_Percentage_2019[i])+"%"
    S2019.append(Seat_Percent)
for i in range(len(Seat_Percentage_2024)):
    Seat_Percent=str(Seat_Percentage_2024[i])+"%"
    S2024.append(Seat_Percent)

Seat_Share_Comparison_dict={"Party":Major_Parties,"Seat Share Percent in 2019":S2019,"Seat Share Percent in 2024":S2024, "Difference":Difference}
Seat_Share_df=pd.DataFrame(Seat_Share_Comparison_dict)
Seat_Share_df.to_excel("Seat Share.xlsx", index=False)
workbook = load_workbook("Seat Share.xlsx")
sheet = workbook.active
sheet.column_dimensions['A'].width = 45
sheet.column_dimensions['B'].width = 26
sheet.column_dimensions['C'].width = 26
sheet.column_dimensions['D'].width = 25
for cell in sheet[1]:  # Assuming headers are in the first row
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")  # White font for contrast
for row_index, row in enumerate(sheet.iter_rows(min_row=2)):  # Start from the second row
    for cell in row:
        fill_color = row_fill1 if row_index % 2 == 0 else row_fill2
        cell.fill = fill_color
workbook.save("Seat Share.xlsx")

fig, ax = plt.subplots()
ax.set_title("Seats secured by Major Parties in 2019 Vs. 2024(in Percentage)",font="Impact",fontsize=15,loc='left',color="white")
ax.set_xlabel("Seats secured in percentage",color="white",fontsize=12)
ax.set_ylabel("Major parties",color="white",fontsize=12)
fig.set_figheight(8)
fig.set_figwidth(16)

fig.patch.set_facecolor("#2F2F2F")  
ax.set_facecolor("#2F2F2F")
ax.tick_params(axis='x', colors='white')
ax.tick_params(axis='y', colors='white') 
ax.grid(color='gray')

ax.plot(MP,Seat_Percentage_2019,marker="o",color="#00796B",linewidth=2)
ax.plot(MP,Seat_Percentage_2024, marker="o", color="#FFB300",linewidth=2)
ax.legend(labels=Labels,fontsize=11)

plt.savefig("Seats secured by Major Parties in 2019 Vs. 2024(in Percentage).png")
plt.show()